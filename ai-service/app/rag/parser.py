"""多格式文档解析：把常见格式统一转为「分页文本」结构。

统一返回 ``[{"page": int, "text": str}]``：
- 分页格式（pdf）逐页返回真实页号；
- 非分页格式（txt/md/csv/json/docx/xlsx）``page`` 固定为 0。

pdf/docx/xlsx 依赖体积较大，采用惰性导入：缺失时抛 ``UnsupportedFormatError``，
避免在没装对应库的环境直接 ImportError 崩掉整个管道。
"""

from __future__ import annotations

from io import BytesIO

# 支持解析的文件扩展名（小写）
SUPPORTED_EXTS: frozenset = frozenset({"txt", "md", "csv", "json", "pdf", "docx", "xlsx"})

# 纯文本类：直接 utf-8 解码即可，无需第三方库
_PLAIN_EXTS = frozenset({"txt", "md", "csv", "json"})


class UnsupportedFormatError(ValueError):
    """文件格式不支持或对应解析依赖未安装。"""


def parse_document(filename: str, content: bytes) -> list[dict]:
    """解析 ``content`` 为 ``[{"page": int, "text": str}]`` 列表。

    ``filename`` 的扩展名决定解析方式；未知扩展名抛 ``UnsupportedFormatError``。
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in _PLAIN_EXTS:
        text = content.decode("utf-8", errors="replace")
        return [{"page": 0, "text": text}]
    if ext == "pdf":
        return _parse_pdf(content)
    if ext == "docx":
        return _parse_docx(content)
    if ext == "xlsx":
        return _parse_xlsx(content)
    raise UnsupportedFormatError(f"不支持的文件格式: {ext or filename}")


def _parse_pdf(content: bytes) -> list[dict]:
    """惰性依赖 pypdf，逐页提取文本并去掉空页。"""
    try:
        from pypdf import PdfReader
    except ImportError:
        raise UnsupportedFormatError("解析 pdf 需要安装 pypdf")
    reader = PdfReader(BytesIO(content))
    pages: list[dict] = []
    for index, page in enumerate(reader.pages):
        text = (page.extract_text() or "").strip()
        if text:
            pages.append({"page": index + 1, "text": text})
    return pages


def _parse_docx(content: bytes) -> list[dict]:
    """惰性依赖 python-docx，把所有段落拼接为一个文本块。"""
    try:
        from docx import Document
    except ImportError:
        raise UnsupportedFormatError("解析 docx 需要安装 python-docx")
    document = Document(BytesIO(content))
    text = "\n".join(paragraph.text for paragraph in document.paragraphs).strip()
    return [{"page": 0, "text": text}]


def _parse_xlsx(content: bytes) -> list[dict]:
    """惰性依赖 openpyxl，按行把非空单元格拼接为文本块。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise UnsupportedFormatError("解析 xlsx 需要安装 openpyxl")
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None and str(cell) != ""]
            if cells:
                lines.append("\t".join(cells))
    text = "\n".join(lines).strip()
    return [{"page": 0, "text": text}]
